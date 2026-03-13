#!/usr/bin/env python3
"""
Example: Convert a Wasabi state JSON export to CSV and Excel.

Demonstrates extracting bucket data from a state snapshot (as produced
by get_wasabi_state.py) and outputting formatted CSV and Excel files.

Requires: openpyxl (pip install openpyxl)
"""

import json
import csv
import argparse
import sys
import os
from typing import Dict, List, Tuple, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def extract_bucket_data(json_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extract and process bucket data from the JSON structure.

    Args:
        json_data: The parsed JSON data containing Wasabi resources

    Returns:
        A list of dictionaries with processed bucket information

    Raises:
        ValueError: If the JSON doesn't contain a 'buckets' section
    """
    buckets_data = []

    if "buckets" not in json_data:
        raise ValueError("Input JSON doesn't contain a 'buckets' section")

    for bucket_name, bucket_info in json_data["buckets"].items():
        bucket_data = {
            "bucket_name": bucket_name,
            "region": bucket_info.get("region", ""),
            "gb_used": bucket_info.get("gb_used", 0),
            "tb_used": bucket_info.get("gb_used", 0) / 1000,
            "object_count": bucket_info.get("object_count", 0),
            "has_lifecycle_policy": "No",
            "category": "",
            "notes": "",
        }

        lifecycle_rules = bucket_info.get("lifecycle-rules", {})
        if lifecycle_rules and isinstance(lifecycle_rules, dict):
            rules = lifecycle_rules.get("Rules", [])
            if rules and any(rule.get("Status") == "Enabled" for rule in rules):
                bucket_data["has_lifecycle_policy"] = "Yes"

        buckets_data.append(bucket_data)

    buckets_data.sort(key=lambda x: x["gb_used"], reverse=True)

    return buckets_data


def get_output_filenames(input_file: str) -> Tuple[str, str]:
    """
    Generate output filenames for CSV and Excel based on input filename.

    Args:
        input_file: Path to the input JSON file

    Returns:
        A tuple containing (csv_filename, excel_filename)
    """
    directory, filename = os.path.split(input_file)
    root = os.path.splitext(filename)[0]
    csv_file = os.path.join(directory, f"{root}.csv")
    xlsx_file = os.path.join(directory, f"{root}.xlsx")

    return csv_file, xlsx_file


def create_csv(bucket_data: List[Dict[str, Any]], output_file: str) -> None:
    """
    Create a CSV file with the extracted bucket data.

    Args:
        bucket_data: List of dictionaries containing bucket information
        output_file: Path where the CSV file will be saved
    """
    csv_fieldnames = [
        "bucket_name",
        "region",
        "gb_used",
        "tb_used",
        "object_count",
        "has_lifecycle_policy",
        "category",
        "notes",
    ]

    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fieldnames)
        writer.writeheader()
        writer.writerows(bucket_data)
    print(f"Successfully created CSV file: {output_file}")


def create_excel(bucket_data: List[Dict[str, Any]], output_file: str) -> None:
    """
    Create an Excel file with formatted bucket data.

    Args:
        bucket_data: List of dictionaries containing bucket information
        output_file: Path where the Excel file will be saved
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bucket Data"

    headers = [
        "Bucket Name",
        "Region",
        "GB Used",
        "TB Used",
        "Object Count",
        "Has Lifecycle Policy",
        "Category",
        "Notes",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color="DDDDDD",
            end_color="DDDDDD",
            fill_type="solid",
        )

    for row_num, data in enumerate(bucket_data, 2):
        ws.cell(row=row_num, column=1).value = data["bucket_name"]
        ws.cell(row=row_num, column=2).value = data["region"]

        gb_cell = ws.cell(row=row_num, column=3)
        gb_cell.value = round(data["gb_used"])
        gb_cell.number_format = "#,##0"

        tb_cell = ws.cell(row=row_num, column=4)
        tb_cell.value = data["tb_used"]
        tb_cell.number_format = "#,##0"

        obj_count_cell = ws.cell(row=row_num, column=5)
        obj_count_cell.value = data["object_count"]
        obj_count_cell.number_format = "#,##0"

        ws.cell(row=row_num, column=6).value = data["has_lifecycle_policy"]
        ws.cell(row=row_num, column=7).value = data["category"]
        ws.cell(row=row_num, column=8).value = data["notes"]

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_file)
    print(f"Successfully created Excel file: {output_file}")


def process_json(input_file: str) -> None:
    """
    Process JSON file and create CSV and Excel outputs.

    Args:
        input_file: Path to the input JSON file
    """
    if not os.path.isfile(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    csv_file, xlsx_file = get_output_filenames(input_file)

    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    bucket_data = extract_bucket_data(data)

    if not bucket_data:
        print("No bucket data found in the JSON file")
        return

    create_csv(bucket_data, csv_file)
    create_excel(bucket_data, xlsx_file)


def main() -> int:
    """Main function to parse arguments and initiate processing."""
    parser = argparse.ArgumentParser(
        description="Convert Wasabi S3 bucket information from JSON to CSV and Excel"
    )
    parser.add_argument("input_file", help="Input JSON file path")

    args = parser.parse_args()

    try:
        process_json(args.input_file)
    except Exception as e:
        print(f"Error: {e}")
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
